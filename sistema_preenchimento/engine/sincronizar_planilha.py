"""
sincronizar_planilha.py
Roda todo dia no final do expediente no servidor Geo (Task Scheduler).

Lê o status atual do Supabase (schema plantio) e escreve de volta nas colunas
AA (Sist. Conser.) / AB (Mapeamento) / AC (Projeto) da aba "Sequencia" da
planilha de Sequência de Plantio configurada em config.json.

Matching feito por (CODIGO, mês-de-MES-DE-PLANTIO, TALHÕES expandidos) —
substitui a lógica antiga de bloco_id que foi descontinuada com o PLANAGRI.

Por que xlwings para escrita: a aba "Sequencia" tem mais de 100 regras de
formatação condicional e listas de validação — abrir e salvar com openpyxl
descarta essas extensões. xlwings usa o Excel real instalado na máquina.
A planilha não pode estar aberta em outra instância durante a sincronização.

Requer: openpyxl, requests, xlwings + Excel instalado no servidor
Config: engine/config.json  (ver config.json.example)
"""

import datetime
import json
import os
import sys

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)
sys.path.insert(0, _SCRIPT_DIR)

from utils import (
    norm_header, strip_accents, redirecionar_stdout, fechar_log,
    achar_header, data_para_iso, rollup_sist_conser, parse_talhoes,
)

# ── Config ────────────────────────────────────────────────────────────────────
# Tenta config.json unificado (modo servidor) primeiro; cai nos JSONs legados
# (supabase_config.json + planilha_config.json) para rodar manualmente na rede.

_cfg = {}
_config_path = os.path.join(_SCRIPT_DIR, 'config.json')
if not os.path.exists(_config_path):
    _config_path = os.path.join(_BASE_DIR, 'config.json')

if os.path.exists(_config_path):
    with open(_config_path, encoding='utf-8') as _f:
        _cfg = json.load(_f)
    SUPABASE_URL   = _cfg['supabase_url'].rstrip('/')
    SUPABASE_KEY   = _cfg['supabase_service_key']
    SOURCE_PLANTIO = _cfg.get('sequencia_path')    # None → busca no diretório
    SENHA_PLANILHA = _cfg.get('senha_planilha') or None
    LOG_DIR        = _cfg.get('log_dir', os.path.join(_BASE_DIR, 'logs'))
    _UNATTENDED    = True    # sem input() quando rodando pelo servidor
else:
    # Modo legado: JSONs separados + busca de planilha no diretório corrente
    _sb_path = os.path.join(_BASE_DIR, 'supabase_config.json')
    if not os.path.exists(_sb_path):
        _sb_path = os.path.join(_SCRIPT_DIR, 'supabase_config.json')
    if not os.path.exists(_sb_path):
        print("ERRO: supabase_config.json não encontrado.")
        sys.exit(1)
    with open(_sb_path, encoding='utf-8') as _f:
        _sb = json.load(_f)
    SUPABASE_URL = _sb['url'].rstrip('/')
    SUPABASE_KEY = _sb['key']

    _pl_path = os.path.join(_BASE_DIR, 'planilha_config.json')
    if not os.path.exists(_pl_path):
        _pl_path = os.path.join(_SCRIPT_DIR, 'planilha_config.json')
    SENHA_PLANILHA = None
    if os.path.exists(_pl_path):
        with open(_pl_path, encoding='utf-8') as _f:
            SENHA_PLANILHA = json.load(_f).get('senha_plantio') or None
    if not SENHA_PLANILHA:
        print("AVISO: senha_plantio não encontrada — planilha com senha de gravação não será salva.\n")

    SOURCE_PLANTIO = None   # vai buscar no diretório corrente
    LOG_DIR        = os.path.join(_BASE_DIR, 'logs')
    _UNATTENDED    = False

os.makedirs(LOG_DIR, exist_ok=True)
_log_fh = redirecionar_stdout(os.path.join(LOG_DIR, f'sincronizar_{datetime.date.today()}.log'))

import glob as _glob    # noqa: E402
import openpyxl         # noqa: E402
import requests         # noqa: E402
import xlwings as xw    # noqa: E402

_SB_HEADERS = {
    'apikey':          SUPABASE_KEY,
    'Authorization':   f'Bearer {SUPABASE_KEY}',
    'Accept-Profile':  'plantio',
    'Content-Profile': 'plantio',
}


def _encerrar(codigo=1):
    fechar_log(_log_fh)
    if not _UNATTENDED:
        input("\nPressione Enter para sair...")
    sys.exit(codigo)


# ── 1. Lê programação do Supabase (paginado) ──────────────────────────────────

print(f"[{datetime.datetime.now():%H:%M:%S}] Lendo programação do Supabase...")

_por_chave = {}   # (cod_faz_int, mes_num, talhao_int) → {sist_conser, mapeamento, projeto}
_offset = 0
_LIMIT  = 1000

while True:
    _r = requests.get(
        f"{SUPABASE_URL}/rest/v1/programacao"
        f"?select=cod_faz,talhao,mes_plantio,sist_conser,mapeamento,projeto"
        f"&limit={_LIMIT}&offset={_offset}",
        headers=_SB_HEADERS,
        timeout=30,
    )
    if not _r.ok:
        print(f"ERRO ao ler programacao: {_r.status_code} {_r.text}")
        _encerrar()
    _batch = _r.json()
    for _row in _batch:
        try:
            _cod = int(_row['cod_faz'])
            _mes = int(_row['mes_plantio']) if _row.get('mes_plantio') else None
            _tal = int(_row['talhao']) if _row.get('talhao') is not None else None
        except (ValueError, TypeError):
            continue
        if _cod and _mes and _tal is not None:
            _por_chave[(_cod, _mes, _tal)] = {
                'sist_conser': _row.get('sist_conser') or '',
                'mapeamento':  _row.get('mapeamento') or 'Não',
                'projeto':     _row.get('projeto') or 'Pendente',
            }
    if len(_batch) < _LIMIT:
        break
    _offset += _LIMIT

print(f"  {len(_por_chave)} talhão-chave(s) carregado(s) do Supabase.")


# ── 2. Rollup por grupo (mesma lógica do frontend) ────────────────────────────

_PROJ_RANK = {'Aguard. Map.': 0, 'Pendente': 1, 'Andamento': 2, 'Ag. Mapa': 3, 'Ok': 4}


def rollup_projeto(valores):
    """Pior rank ganha — igual ao getFazEtapa do web."""
    pior = 'Ok'
    for v in valores:
        p = v or 'Pendente'
        if _PROJ_RANK.get(p, 1) < _PROJ_RANK.get(pior, 4):
            pior = p
    return pior


def rollup_grupo(talhoes_dados):
    sc   = rollup_sist_conser([d['sist_conser'] for d in talhoes_dados])
    mapa = 'Sim' if all(d['mapeamento'] == 'Sim' for d in talhoes_dados) else 'Não'
    proj = rollup_projeto([d['projeto'] for d in talhoes_dados])
    return sc, mapa, proj


# ── 3. Localiza a planilha de Sequência de Plantio ───────────────────────────

print(f"[{datetime.datetime.now():%H:%M:%S}] Localizando planilha...")

if SOURCE_PLANTIO:
    if not os.path.exists(SOURCE_PLANTIO):
        print(f"ERRO: planilha não encontrada: {SOURCE_PLANTIO}")
        _encerrar()
    sheet_seq_name = None
    _wb_check = openpyxl.load_workbook(SOURCE_PLANTIO, read_only=True, data_only=True)
    sheet_seq_name = next((n for n in _wb_check.sheetnames if norm_header(n) == 'SEQUENCIA'), None)
    _wb_check.close()
    if not sheet_seq_name:
        print(f"ERRO: aba 'Sequencia' não encontrada em {SOURCE_PLANTIO}")
        _encerrar()
else:
    # Modo legado: busca no diretório corrente
    _cwd = os.getcwd()
    _candidates = sorted(f for f in _glob.glob(os.path.join(_cwd, '*.xlsx'))
                         if not os.path.basename(f).startswith('~$'))
    SOURCE_PLANTIO = None
    sheet_seq_name = None
    for _c in _candidates:
        _wb_c = openpyxl.load_workbook(_c, read_only=True, data_only=True)
        _s = next((n for n in _wb_c.sheetnames if norm_header(n) == 'SEQUENCIA'), None)
        _wb_c.close()
        if _s:
            SOURCE_PLANTIO = os.path.abspath(_c)
            sheet_seq_name = _s
            break
    if not SOURCE_PLANTIO:
        print(f"ERRO: nenhum .xlsx com aba 'Sequencia' encontrado em {_cwd}")
        _encerrar()

print(f"  Planilha: {SOURCE_PLANTIO}")


# ── 4. Lê linhas (openpyxl, somente leitura) e faz o match ───────────────────

_wb_ro = openpyxl.load_workbook(SOURCE_PLANTIO, data_only=True)
_ws_ro = _wb_ro[sheet_seq_name]

_LIMITE_COL = openpyxl.utils.column_index_from_string('AC')
_header_row, _hmap = achar_header(
    _ws_ro,
    ['MES DE PLANTIO', 'CODIGO', 'TALHOES'],
    max_row=10,
    max_col=_LIMITE_COL,
)
if _header_row is None:
    print("ERRO: cabeçalho (Mês de Plantio / Código / Talhões) não encontrado.")
    _wb_ro.close()
    _encerrar()

_idx_mes = _hmap.get('MES DE PLANTIO')
_idx_cod = _hmap.get('CODIGO')
_idx_tal = _hmap.get('TALHOES')


def _mes_raw_para_numero(v):
    """Extrai o mês (1-12) de uma célula que pode ser datetime, date ou string."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.month
    if isinstance(v, str):
        s = v.strip()
        _MESES = {
            'JANEIRO':1,'FEVEREIRO':2,'MARCO':3,'ABRIL':4,'MAIO':5,'JUNHO':6,
            'JULHO':7,'AGOSTO':8,'SETEMBRO':9,'OUTUBRO':10,'NOVEMBRO':11,'DEZEMBRO':12,
        }
        num = _MESES.get(strip_accents(s).upper())
        if num:
            return num
        try:
            return datetime.date.fromisoformat(s[:10]).month
        except ValueError:
            pass
    return None


linhas_para_escrever = []
n_sem_dados = 0

for _i, _row in enumerate(
    _ws_ro.iter_rows(
        min_row=_header_row + 1,
        max_row=_ws_ro.max_row,
        max_col=_LIMITE_COL,
        values_only=True,
    ),
    start=_header_row + 1,
):
    _cod_raw = _row[_idx_cod] if _idx_cod < len(_row) else None
    _tal_raw = _row[_idx_tal] if _idx_tal < len(_row) else None
    _mes_raw = _row[_idx_mes] if _idx_mes < len(_row) else None

    if _cod_raw is None or _tal_raw is None:
        continue
    try:
        _cod_int = int(_cod_raw)
    except (ValueError, TypeError):
        continue

    _mes_num  = _mes_raw_para_numero(_mes_raw)
    _talhoes  = parse_talhoes(_tal_raw)

    if not _talhoes or _mes_num is None:
        n_sem_dados += 1
        continue

    _matches = [
        _por_chave[(_cod_int, _mes_num, t)]
        for t in _talhoes
        if (_cod_int, _mes_num, t) in _por_chave
    ]

    if not _matches:
        continue

    sc, mapa, proj = rollup_grupo(_matches)
    linhas_para_escrever.append((_i, sc, mapa, proj))

_wb_ro.close()

print(f"  {len(linhas_para_escrever)} linha(s) a atualizar | {n_sem_dados} sem match de dados.")

if not linhas_para_escrever:
    print("Nada para escrever na planilha.")
    fechar_log(_log_fh)
    sys.exit(0)


# ── 5. Escreve via Excel real (preserva formatação condicional/validação) ─────

COL_AA = 27   # Sist. Conser.
COL_AB = 28   # Mapeamento
COL_AC = 29   # Projeto / Mapa

print(f"[{datetime.datetime.now():%H:%M:%S}] Abrindo Excel para escrita...")

_tmp_path = SOURCE_PLANTIO + '.sync_tmp.xlsx'
if os.path.exists(_tmp_path):
    os.remove(_tmp_path)

_app = xw.App(visible=False)
_app.display_alerts = False
try:
    _wb = _app.books.open(
        SOURCE_PLANTIO,
        password=SENHA_PLANILHA,
        write_res_password=SENHA_PLANILHA,
    )
    print(f"  ReadOnly={_wb.api.ReadOnly}")
    _ws = _wb.sheets[sheet_seq_name]

    for _linha, _sc, _mapa, _proj in linhas_para_escrever:
        _ws.range((_linha, COL_AA)).value = _sc
        _ws.range((_linha, COL_AB)).value = _mapa
        _ws.range((_linha, COL_AC)).value = _proj

    # Carimbo de data na linha 3, coluna AC
    _ws.range((3, COL_AC)).value = f"Atualizado em: {datetime.date.today().strftime('%d/%m/%Y')}"

    # Verificação rápida: lê de volta a primeira linha escrita
    _l0, _sc0, _ma0, _pr0 = linhas_para_escrever[0]
    _check_sc = _ws.range((_l0, COL_AA)).value
    _check_ma = _ws.range((_l0, COL_AB)).value
    _check_pr = _ws.range((_l0, COL_AC)).value
    print(f"  Verificação linha {_l0}: AA={_check_sc!r} AB={_check_ma!r} AC={_check_pr!r}")

    _wb.save(_tmp_path)
finally:
    _wb.close()
    _app.quit()

_orig_size = os.path.getsize(SOURCE_PLANTIO)
_tmp_size  = os.path.getsize(_tmp_path)
print(f"  tmp: {_tmp_size} bytes | original: {_orig_size} bytes")

try:
    os.replace(_tmp_path, SOURCE_PLANTIO)
    print("  Arquivo substituído (os.replace).")
except OSError as _e:
    import shutil
    print(f"  os.replace falhou ({_e}), usando shutil.copy2...")
    shutil.copy2(_tmp_path, SOURCE_PLANTIO)
    os.remove(_tmp_path)

print(f"\n✓ Sincronização concluída: {len(linhas_para_escrever)} linha(s) atualizadas.")
fechar_log(_log_fh)
