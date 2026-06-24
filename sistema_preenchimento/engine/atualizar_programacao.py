"""
atualizar_programacao.py
Uso: python atualizar_programacao.py  OU  duplo clique no ATUALIZAR.bat

Estrutura esperada:
  base_plantio/*.xlsx      ← Sequência de Plantio (aba "Sequencia", demanda: bloco de talhões + Mês de Plantio)
  base_preparo/*.xlsx      ← Planilha de Preparo (aba "CONSERVAÇÃO", Sist. Conser. por talhão individual)
  base_fazendas/*.xlsx     ← base mestre de talhões (área por COD FAZ + TALHÃO)
  supabase_config.json     ← { "url": "...", "key": "sb_publishable_..." }

Lê a aba "Sequencia" (uma linha por bloco, talhões agrupados em lista/faixa,
ex: "1,2,3" ou "1 AO 4"), explode cada linha em um registro por talhão, busca
o Sist. Conser. de cada talhão na aba "CONSERVAÇÃO" da planilha de Preparo e a
área na base_fazendas, e faz upsert em `programacao` no Supabase — preservando
MAPEAMENTO/PROJETO já preenchidos para LAYERs existentes.

bloco_id: a coluna ORDEM da Sequência não é usada pelo time (fica vazia), então
o agrupamento por bloco é reconstruído a partir do conteúdo de cada linha
(COD FAZ + MÊS DE PLANTIO + texto bruto de TALHÕES) — estável a reordenação de
linhas, mas não sobrevive a uma edição desses campos entre sincronizações.
"""

import os
import sys

# ── Utilitários compartilhados ────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)   # sistema_preenchimento/
sys.path.insert(0, _SCRIPT_DIR)
from utils import (layer_to_str, norm_header, parse_talhoes, redirecionar_stdout, fechar_log,
                    achar_header, data_para_iso, rollup_sist_conser, SIST_CONSER_PRECISA_MAPEAMENTO,
                    normaliza_sist_conser)

_log_fh = redirecionar_stdout(os.path.join(_BASE_DIR, 'logs', 'atualizar.log'))

import datetime
import glob as _glob
import json

import pandas as pd
import requests
import openpyxl

# ── Carrega configurações ─────────────────────────────────────────────────
_cfg_path = os.path.join(_BASE_DIR, 'config.json')
try:
    with open(_cfg_path, 'r', encoding='utf-8') as _f:
        _cfg = json.load(_f)
except Exception:
    _cfg = {}

CODFAZ_EXCLUIR_PREFIXO = _cfg.get('codfaz_excluir_prefixo', '20')

_config_path = os.path.join(_BASE_DIR, 'supabase_config.json')
if not os.path.exists(_config_path):
    print(f"ERRO: Arquivo nao encontrado → {_config_path}")
    print("  Crie esse arquivo com: { \"url\": \"https://xxxx.supabase.co\", \"key\": \"sb_publishable_...\" }")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

with open(_config_path, 'r', encoding='utf-8') as _f:
    _sb_cfg = json.load(_f)

SUPABASE_URL = _sb_cfg['url'].rstrip('/')
SUPABASE_KEY = _sb_cfg['key']
# Projeto compartilhado com o project-preparo — tabelas do Plantio isoladas no
# schema "plantio" (ver migration), por isso Accept-Profile/Content-Profile.
SB_HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
    'Accept-Profile': 'plantio',
    'Content-Profile': 'plantio',
}

os.chdir(_BASE_DIR)

# ── 1. Ler valores existentes no Supabase (preservar preenchimento) ──────
print("Lendo programação existente no Supabase...")
_res = requests.get(f"{SUPABASE_URL}/rest/v1/programacao?select=layer,mapeamento,projeto", headers=SB_HEADERS)
if not _res.ok:
    print(f"ERRO ao ler programacao: {_res.status_code} {_res.text}")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
preserved = {}   # layer_str → (mapeamento, projeto)
for row in _res.json():
    layer = layer_to_str(row.get('layer'))
    if layer:
        preserved[layer] = (row.get('mapeamento') or 'Não', row.get('projeto') or 'Pendente')
print(f"  {len(preserved)} linhas existentes carregadas.\n")

# ── 2. Ler base_fazendas (área por COD FAZ + TALHÃO) ──────────────────────
print("Verificando base fazendas...")
_base_faz_files = _glob.glob("base_fazendas/*.xls*")
layer_ha = {}   # (cod_faz, talhao) → area_ha
if not _base_faz_files:
    print("  AVISO: Nenhum arquivo em base_fazendas/ — área de cada talhão ficará 0.\n")
else:
    SOURCE_BASE = _base_faz_files[0]
    print(f"  Base fazendas: {SOURCE_BASE}")
    df_base = pd.read_excel(SOURCE_BASE, engine='openpyxl')
    df_base.columns = [norm_header(c) for c in df_base.columns]
    col_cod = next((c for c in ('CODIGO', 'COD FAZ', 'SECAO') if c in df_base.columns), None)
    col_tal = next((c for c in ('TALHAO', 'TALHOES') if c in df_base.columns), None)
    col_area = next((c for c in ('AREA_PROD', 'AREA_HA', 'HA', 'AREA') if c in df_base.columns), None)
    if not col_cod or not col_tal or not col_area:
        print(f"  AVISO: colunas esperadas (COD FAZ/TALHAO/AREA) não encontradas em {SOURCE_BASE} — área ficará 0.\n")
    else:
        df_base[col_cod] = pd.to_numeric(df_base[col_cod], errors='coerce')
        df_base[col_tal] = pd.to_numeric(df_base[col_tal], errors='coerce')
        df_base[col_area] = pd.to_numeric(df_base[col_area], errors='coerce')
        df_base = df_base.dropna(subset=[col_cod, col_tal])
        for _, r in df_base.iterrows():
            try:
                key = (int(r[col_cod]), int(r[col_tal]))
                layer_ha[key] = round(float(r[col_area] or 0), 2)
            except (ValueError, TypeError):
                pass
        print(f"  {len(layer_ha)} talhões na base fazendas.\n")

# ── 3. Ler aba CONSERVAÇÃO de todas as planilhas de Preparo (Sist. Conser. por talhão) ──
# Pode haver mais de um arquivo aqui (ex: ano-safra atual + histórico de ano-safra
# anterior, divididos em arquivos separados) — lê todos e funde num único lookup.
print("Lendo planilha(s) de preparo (aba CONSERVAÇÃO)...")
_preparo_found = _glob.glob("base_preparo/*.xlsx")
if not _preparo_found:
    print("ERRO: Nenhum arquivo .xlsx encontrado em base_preparo/")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

sist_conser_por_layer = {}   # (cod_faz, talhao) → sist_conser
for SOURCE_PREPARO in _preparo_found:
    print(f"  Planilha de preparo: {SOURCE_PREPARO}")
    wb_preparo = openpyxl.load_workbook(SOURCE_PREPARO, data_only=True)
    sheet_cons = next((n for n in wb_preparo.sheetnames if norm_header(n) == 'CONSERVACAO'), None)
    if not sheet_cons:
        print(f"  ⚠  Aba 'CONSERVAÇÃO' não encontrada em {SOURCE_PREPARO} (abas: {wb_preparo.sheetnames}) — ignorada.")
        continue
    ws_cons = wb_preparo[sheet_cons]

    header_row, hmap = achar_header(ws_cons, ['SECAO', 'TALHAO', 'SISTEMA DE CONSERVACAO'])
    if header_row is None:
        print(f"  ⚠  Cabeçalho (SEÇÃO/TALHÃO/SISTEMA DE CONSERVAÇÃO) não encontrado em {SOURCE_PREPARO} — ignorada.")
        continue

    idx_cod_cons = hmap['SECAO']
    idx_tal_cons = hmap['TALHAO']
    idx_sc       = hmap['SISTEMA DE CONSERVACAO']

    n_antes = len(sist_conser_por_layer)
    for row in ws_cons.iter_rows(min_row=header_row + 1, max_row=ws_cons.max_row, values_only=True):
        cod_raw = row[idx_cod_cons] if idx_cod_cons < len(row) else None
        tal_raw = row[idx_tal_cons] if idx_tal_cons < len(row) else None
        sc_raw  = row[idx_sc] if idx_sc < len(row) else None
        if cod_raw is None or tal_raw is None:
            continue
        try:
            cod_faz = int(cod_raw)
            talhao  = int(tal_raw)
        except (ValueError, TypeError):
            continue
        sist_conser_por_layer[(cod_faz, talhao)] = normaliza_sist_conser(sc_raw)
    print(f"    {len(sist_conser_por_layer) - n_antes} talhão(ões) novo(s)/atualizado(s).")

print(f"  {len(sist_conser_por_layer)} talhão(ões) com Sist. Conser. no total (todas as planilhas de preparo).\n")

# ── 4. Ler aba Sequencia da planilha de Plantio (demanda, blocos de talhões) ──
print("Lendo planilha de Sequência de Plantio (aba Sequencia)...")
_plantio_found = _glob.glob("base_plantio/*.xlsx")
if not _plantio_found:
    print("ERRO: Nenhum arquivo .xlsx encontrado em base_plantio/")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
SOURCE_PLANTIO = _plantio_found[0]
print(f"  Planilha de plantio: {SOURCE_PLANTIO}")

wb_plantio = openpyxl.load_workbook(SOURCE_PLANTIO, data_only=True)
sheet_seq = next((n for n in wb_plantio.sheetnames if norm_header(n) == 'SEQUENCIA'), None)
if not sheet_seq:
    print(f"ERRO: Aba 'Sequencia' não encontrada. Abas disponíveis: {wb_plantio.sheetnames}")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
ws_seq = wb_plantio[sheet_seq]

# Colunas além de AB (bloco "MUDA") são ignoradas — fora de escopo. Também repetem
# nomes de coluna (Código/Seção/Talhão), então a busca de cabeçalho precisa se
# limitar a A:AB pra não capturar os índices errados.
LIMITE_COL = openpyxl.utils.column_index_from_string('AB')

header_row, hmap = achar_header(ws_seq, ['MES DE PLANTIO', 'CODIGO', 'SECAO', 'TALHOES'], max_row=10, max_col=LIMITE_COL)
if header_row is None:
    print(f"ERRO: Cabeçalho (Mês de Plantio/Código/Seção/Talhões) não encontrado na aba '{sheet_seq}'.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

idx_mes = hmap['MES DE PLANTIO']
idx_cic = hmap.get('CICLO')
idx_cod = hmap['CODIGO']
idx_sec = hmap['SECAO']
idx_tal = hmap['TALHOES']
idx_amb = hmap.get('AMBIENTE')

# ── 5. Explode cada linha (bloco de talhões) em um registro por talhão ────
exploded = []   # {cod_faz, fazenda, talhao, mes_plantio, ciclo, ambiente, bloco_id}
n_linhas = n_ignoradas = n_falha_talhao = n_admin = 0
for row in ws_seq.iter_rows(min_row=header_row + 1, max_row=ws_seq.max_row, max_col=LIMITE_COL, values_only=True):
    cod_raw = row[idx_cod] if idx_cod < len(row) else None
    sec_raw = row[idx_sec] if idx_sec < len(row) else None
    tal_raw = row[idx_tal] if idx_tal < len(row) else None
    if cod_raw is None or sec_raw is None or tal_raw is None:
        if cod_raw is not None or sec_raw is not None or tal_raw is not None:
            n_ignoradas += 1
        continue
    try:
        cod_faz = int(cod_raw)
    except (ValueError, TypeError):
        n_ignoradas += 1
        continue
    if str(cod_faz).startswith(CODFAZ_EXCLUIR_PREFIXO):
        n_admin += 1
        continue

    talhoes = parse_talhoes(tal_raw)
    if not talhoes:
        print(f"  ⚠  COD FAZ {cod_faz}: não foi possível interpretar TALHÕES = {tal_raw!r} — linha ignorada.")
        n_falha_talhao += 1
        continue

    n_linhas += 1
    fazenda     = str(sec_raw).strip()
    mes_plantio = data_para_iso(row[idx_mes]) if idx_mes < len(row) else None
    ciclo       = str(row[idx_cic]).strip() if idx_cic is not None and idx_cic < len(row) and row[idx_cic] is not None else ''
    ambiente    = str(row[idx_amb]).strip() if idx_amb is not None and idx_amb < len(row) and row[idx_amb] else ''
    bloco_id    = f"{cod_faz}|{mes_plantio}|{str(tal_raw).strip()}"

    for talhao in talhoes:
        exploded.append({
            'cod_faz': cod_faz, 'fazenda': fazenda, 'talhao': talhao,
            'mes_plantio': mes_plantio, 'ciclo': ciclo, 'ambiente': ambiente,
            'bloco_id': bloco_id,
        })

print(f"  {n_linhas} linha(s)/bloco(s) de demanda, {len(exploded)} talhão(ões) explodido(s).")
if n_admin:        print(f"  Filtro administrativo (COD FAZ {CODFAZ_EXCLUIR_PREFIXO}x): {n_admin} linha(s) excluída(s).")
if n_ignoradas:     print(f"  {n_ignoradas} linha(s) parcial(is) ignorada(s) (faltava CÓDIGO/SEÇÃO/TALHÕES).")
if n_falha_talhao:  print(f"  {n_falha_talhao} linha(s) com TALHÕES não interpretável.")
print()

# ── 6. Monta registros por LAYER (último ganha em caso de talhão duplicado) ──
print("Montando registros para o Supabase...")
sem_area = 0
sem_sist_conser = 0
por_layer = {}
for r in exploded:
    layer_val = int(f"{r['cod_faz']}{r['talhao']:03d}")
    area_ha = layer_ha.get((r['cod_faz'], r['talhao']), 0)
    if not area_ha:
        sem_area += 1
    sist_conser = sist_conser_por_layer.get((r['cod_faz'], r['talhao']))
    if sist_conser is None:
        sist_conser = ''   # ainda sem registro na base de preparo (mesmo valor canônico de "vazio")
        sem_sist_conser += 1
    periodo_op = None
    if r['mes_plantio']:
        periodo_op = int(r['mes_plantio'].split('-')[1])
    por_layer[layer_val] = {
        'layer': layer_val,
        'periodo_op': periodo_op,
        'cod_faz': r['cod_faz'],
        'fazenda': r['fazenda'],
        'talhao': r['talhao'],
        'area_ha': area_ha,
        'mes_plantio': r['mes_plantio'],
        'ciclo': r['ciclo'],
        'ambiente': r['ambiente'],
        'sist_conser': sist_conser,
        'bloco_id': r['bloco_id'],
    }
n_dup = len(exploded) - len(por_layer)
if n_dup:
    print(f"  {n_dup} talhão(ões) duplicado(s) entre linhas — mantida a última ocorrência.")
if sem_area:
    print(f"  ⚠  {sem_area} talhão(ões) sem área encontrada na base_fazendas (area_ha = 0).")
if sem_sist_conser:
    print(f"  ⚠  {sem_sist_conser} talhão(ões) ainda sem registro na base de preparo (sist_conser = '', preparo provavelmente não chegou na etapa de conservação).")

# ── Status inicial do Projeto, por bloco (mesma regra do formulario.html) ──
# 'Aguard. Map.' quando o rollup do bloco precisa de mapeamento e o Mapeamento
# ainda não é 'Sim'; senão 'Pendente'. Só corrige layers que ainda estão numa
# dessas duas situações "não iniciadas" — nunca sobrescreve Andamento/Ok.
_bloco_sist_conser = {}
for rec in por_layer.values():
    _bloco_sist_conser.setdefault(rec['bloco_id'], []).append(rec['sist_conser'])
_bloco_rollup = {bid: rollup_sist_conser(vals) for bid, vals in _bloco_sist_conser.items()}


def _status_inicial(bloco_id, mapeamento):
    sc = _bloco_rollup.get(bloco_id, '')
    if sc in SIST_CONSER_PRECISA_MAPEAMENTO and mapeamento != 'Sim':
        return 'Aguard. Map.'
    return 'Pendente'


novos = 0
corrigidos = 0
mapeamento_corrigido = 0
prog_rows = []
for layer_val, rec in por_layer.items():
    ly_str = layer_to_str(layer_val)
    sc_bloco = _bloco_rollup.get(rec['bloco_id'], '')
    # Base larga/vazio não tem o que mapear — Mapeamento já nasce/fica 'Sim' (ver SETUP.md).
    sem_mapeamento_pendente = sc_bloco not in SIST_CONSER_PRECISA_MAPEAMENTO
    if ly_str in preserved:
        mapeamento, projeto = preserved[ly_str]
        if sem_mapeamento_pendente and mapeamento != 'Sim':
            mapeamento = 'Sim'
            mapeamento_corrigido += 1
        if projeto in ('Pendente', 'Aguard. Map.'):
            novo_projeto = _status_inicial(rec['bloco_id'], mapeamento)
            if novo_projeto != projeto:
                corrigidos += 1
            projeto = novo_projeto
    else:
        mapeamento = 'Sim' if sem_mapeamento_pendente else 'Não'
        projeto = _status_inicial(rec['bloco_id'], mapeamento)
        novos += 1
    prog_rows.append(dict(rec, mapeamento=mapeamento, projeto=projeto))

print(f"\nEnviando {len(prog_rows)} linhas (upsert por LAYER)...")
HEADERS_UPSERT = dict(SB_HEADERS, Prefer='resolution=merge-duplicates,return=minimal')
BATCH = 500
for i in range(0, len(prog_rows), BATCH):
    chunk = prog_rows[i:i+BATCH]
    res = requests.post(f"{SUPABASE_URL}/rest/v1/programacao", headers=HEADERS_UPSERT, json=chunk)
    if not res.ok:
        print(f"ERRO ao enviar lote {i}-{i+len(chunk)}: {res.status_code} {res.text}")
        fechar_log(_log_fh)
        input("\nPressione Enter para sair...")
        sys.exit(1)
    print(f"  {min(i+BATCH, len(prog_rows))}/{len(prog_rows)}")

print(f"\n  Upsert concluído. Novas linhas: {novos}. Status corrigido (Pendente ⇄ Aguard. Map.): {corrigidos}. "
      f"Mapeamento auto-corrigido p/ Sim (Base larga/vazio): {mapeamento_corrigido}.\n")

# ── Aviso: LAYERs com preenchimento removidos desta atualização ───────────
layers_novos_str = {layer_to_str(r['layer']) for r in prog_rows}
layers_com_preenchimento = {ly for ly, vals in preserved.items() if any(str(v).strip() for v in vals if v not in ('Não', 'Pendente'))}
removidos = layers_com_preenchimento - layers_novos_str
if removidos:
    print(f"  ⚠  ATENÇÃO: {len(removidos)} LAYER(s) preenchidos não estão na nova base:")
    for ly in sorted(removidos)[:10]:
        mapeamento, projeto = preserved[ly]
        print(f"     LAYER {ly}: mapeamento={mapeamento} | projeto={projeto}")
    if len(removidos) > 10:
        print(f"     ... e mais {len(removidos)-10}")
    print("     Esses dados continuam no Supabase — apenas saíram da base atual.\n")

# ── Resumo ────────────────────────────────────────────────────────────────
print(f"{'='*50}")
print(f"  Atualizacao concluida!")
print(f"  Blocos de demanda : {n_linhas}")
print(f"  Talhões explodidos: {len(exploded)}")
print(f"  Registros enviados: {len(prog_rows)}")
print(f"  Sem área (HA=0)   : {sem_area}")
print(f"  Sem Sist. Conser. : {sem_sist_conser}")
print(f"  Preservados       : {len(prog_rows) - novos}")
print(f"{'='*50}")

fechar_log(_log_fh)
input("\nPressione Enter para fechar...")
